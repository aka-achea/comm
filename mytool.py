#!/usr/bin/python3
#coding:utf-8
#tested in win
__version__ = 20191208


'''
Module for misc tool
'''


import time
import sys
import pyautogui as auto
import win32con
import win32clipboard as wincld
import openpyxl
from collections import namedtuple

class myTimer:
    def __init__(self, func=time.perf_counter):
        self.elapsed = 0.0
        self._func = func
        self._start = None

    def start(self):
        if self._start is not None:
            raise RuntimeError('Already started')
        self._start = self._func()

    def stop(self):
        if self._start is None:
            raise RuntimeError('Not started')
        end = self._func()
        self.elapsed += end - self._start
        self._start = None

    def reset(self):
        self.elapsed = 0.0

    @property
    def running(self):
        return self._start is not None

    def __enter__(self):
        self.start()
        return self

    def __exit__(self, *args):
        self.stop()


def mywait(n):
    '''Wait for n seconds'''
    import time
    for i in range(n):
        msg = f'Wait {n-i}'
        print(msg,end='\r')
        time.sleep(1)
        # sys.stdout.write('\x08'*len(msg)+' '*len(msg)+'\x08'*len(msg))
        print(' '*len(msg),end='\r')


def mybar(dots,total):
    bar = '▇'*dots+'--'*(total-dots)
    percentage = f'{dots}/{total}'
    msg = bar+' '+percentage 
    print(msg, end='\r')


def combinekey(a,b):
    '''press two key'''
    auto.keyDown(a)
    auto.keyDown(b)
    auto.keyUp(a)
    auto.keyUp(b)


def capture(img,trys=10,confidence=0.9):  
    '''Find botton location'''  
    trytime = 1
    while trytime < trys:
        try:
            button = auto.locateCenterOnScreen(img,grayscale=True,confidence=confidence)
            if button == None:
                continue
            time.sleep(1)
            break            
        # except TypeError:
        #     time.sleep(15)
        #     trytime += 1
        #     print(f'try to locate {img} {str(trytime)}')
        #     continue
        except OSError as e:
            print(e)
            return e
    else:
        print(f"Max tries reach, not able to locate option {img}")
        button = None
        # return f"Max tries reach, not able to locate option {img}"
        # raise(f'Find no {img}')
    # print(button)
    return button


def clickbutton(img,**args):
    '''click button'''
    button = capture(img,**args)
    if isinstance(button,tuple):
        auto.click(button)
        return True
    else:
        return False


def get_text_clipboard():
    '''Get text from clipboard'''
    wincld.OpenClipboard()
    try:
        text_result = wincld.GetClipboardData(win32con.CF_UNICODETEXT)
    except TypeError:
        return None
    wincld.EmptyClipboard()
    wincld.CloseClipboard()
    return text_result


def build_namedict(xls):
    '''Automatically build named dict from 1st sheet of xls'''
    wb = openpyxl.load_workbook(xls)
    sheets = wb.sheetnames
    sheet = wb[sheets[0]]
    properties = []
    c = 0
    while True:
        c += 1
        if v := sheet.cell(row=1,column=c).value:
            properties.append(v.replace(' ',''))
        else:
            break
    # for a in properties: print(a)
    print(properties)
    tagdic = namedtuple(sheets[0],properties)
    namedict = []
    for r in range(2,sheet.max_row):
            
        tagdic()



if __name__ == "__main__":
    # f = r'M:\GH\a.txt'
    # remove_emptyline_file(f)
    xls = r'E:\UT\iam.xlsx'
    build_namedict(xls)